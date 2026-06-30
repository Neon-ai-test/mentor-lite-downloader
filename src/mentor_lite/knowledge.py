from __future__ import annotations

import base64
import csv
import hashlib
import re
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from mentor_lite.models import KnowledgePoint

STANDARD_FIELDS: list[dict[str, object]] = [
    {
        "key": "chapter",
        "label": "章节",
        "required": True,
        "description": "例如：第一章 特殊平行四边形。空白单元格会沿用上一行章节。",
    },
    {
        "key": "group",
        "label": "一级知识点",
        "required": True,
        "description": "例如：菱形的性质与判定。空白单元格会沿用上一行一级知识点。",
    },
    {
        "key": "name",
        "label": "二级知识点",
        "required": True,
        "description": "真正用于创建粗筛任务和匹配视频的知识点名称。",
    },
]

REQUIRED_FIELDS = {str(field["key"]) for field in STANDARD_FIELDS if field.get("required")}
SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xlsm"}
FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "chapter": ("chapter", "章节", "章", "单元", "模块", "教材章节", "目录"),
    "group": (
        "group",
        "level_1",
        "一级知识点",
        "知识点组",
        "知识模块",
        "知识单元",
        "小节",
    ),
    "name": (
        "name",
        "level_2",
        "knowledge_point",
        "知识点",
        "知识点名称",
        "二级知识点",
        "考点",
        "标题",
    ),
}

SPACE_RE = re.compile(r"\s+")
SAFE_NAME_RE = re.compile(r"[^0-9a-zA-Z._\-\u4e00-\u9fff]+")
ENUM_RE = re.compile(r"^\s*\d+(?:[.、．]\d+)*[.、．\s]*")


@dataclass(frozen=True, slots=True)
class WorksheetData:
    sheet_name: str
    header_row: int
    headers: list[str]
    rows: list[dict[str, str]]


def standard_fields() -> list[dict[str, object]]:
    return [dict(field) for field in STANDARD_FIELDS]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return SPACE_RE.sub(" ", str(value)).strip()


def normalize_key(value: object) -> str:
    return re.sub(r"[^0-9a-zA-Z\u4e00-\u9fff]+", "", normalize_text(value)).lower()


def clean_heading(value: object) -> str:
    return ENUM_RE.sub("", normalize_text(value)).strip()


def stable_id(point: dict[str, object]) -> str:
    basis = "|".join(
        normalize_text(point.get(key))
        for key in ("subject", "stage", "grade", "textbook", "chapter", "group", "name")
    )
    digest = hashlib.sha1(basis.encode("utf-8")).hexdigest()[:12]
    return f"kp_{digest}"


def make_knowledge_point(data: dict[str, Any], point_id: str | None = None) -> KnowledgePoint:
    name = normalize_text(data.get("name"))
    aliases: list[str] = []
    raw_aliases = data.get("aliases", "")
    if isinstance(raw_aliases, list):
        alias_values = raw_aliases
    else:
        alias_values = re.split(r"[\n,，、；;]+", str(raw_aliases or ""))
    for raw in alias_values:
        alias = normalize_text(raw)
        if alias and alias not in aliases:
            aliases.append(alias)
    if name and name not in aliases:
        aliases.insert(0, name)

    payload: dict[str, object] = {
        "subject": normalize_text(data.get("subject")),
        "stage": normalize_text(data.get("stage")),
        "grade": normalize_text(data.get("grade")),
        "textbook": normalize_text(data.get("textbook")),
        "chapter": clean_heading(data.get("chapter")),
        "group": clean_heading(data.get("group")),
        "name": name,
    }
    if not payload["subject"] or not payload["name"]:
        raise ValueError("知识点必须包含学科和名称")
    return KnowledgePoint(
        id=point_id or stable_id(payload),
        subject=str(payload["subject"]),
        stage=str(payload["stage"]),
        grade=str(payload["grade"]),
        textbook=str(payload["textbook"]),
        chapter=str(payload["chapter"]),
        group=str(payload["group"]),
        name=str(payload["name"]),
        description=normalize_text(data.get("description")),
        aliases=aliases,
    )


def save_upload(upload_dir: Path, filename: str, content_base64: str) -> dict[str, str]:
    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError("仅支持 .xlsx、.xlsm 或 .csv 文件；旧版 .xls 请先另存为 .xlsx")
    safe_name = safe_filename(filename)
    upload_id = uuid.uuid4().hex
    upload_dir.mkdir(parents=True, exist_ok=True)
    file_path = upload_dir / f"{upload_id}_{safe_name}"
    try:
        raw = base64.b64decode(content_base64, validate=True)
    except ValueError as exc:
        raise ValueError("上传内容不是有效的 Base64 文件数据") from exc
    if not raw:
        raise ValueError("上传文件为空")
    file_path.write_bytes(raw)
    return {"upload_id": upload_id, "filename": safe_name, "file_path": str(file_path)}


def safe_filename(filename: str) -> str:
    name = Path(filename).name or "knowledge.xlsx"
    safe = SAFE_NAME_RE.sub("_", name).strip("._")
    return safe or "knowledge.xlsx"


def preview_upload(
    file_path: Path,
    *,
    sheet_name: str | None = None,
    header_row: int | None = None,
) -> dict[str, object]:
    worksheet = read_worksheet(file_path, sheet_name=sheet_name, header_row=header_row)
    return {
        "filename": file_path.name,
        "sheet_name": worksheet.sheet_name,
        "header_row": worksheet.header_row,
        "columns": worksheet.headers,
        "sample_rows": worksheet.rows[:8],
        "standard_fields": standard_fields(),
        "suggested_mapping": suggest_mapping(worksheet.headers),
    }


def read_table(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    worksheet = read_worksheet(path)
    return worksheet.headers, worksheet.rows


def read_worksheet(
    file_path: Path,
    *,
    sheet_name: str | None = None,
    header_row: int | None = None,
) -> WorksheetData:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        matrix = read_csv_matrix(file_path)
        selected_sheet = "CSV"
    elif suffix in {".xlsx", ".xlsm"}:
        matrix, selected_sheet = read_excel_matrix(file_path, sheet_name)
    else:
        raise ValueError("仅支持 .xlsx、.xlsm 或 .csv 文件；旧版 .xls 请先另存为 .xlsx")
    if not matrix:
        raise ValueError("文件中没有可读取的数据")
    header_index = header_row - 1 if header_row else detect_header_row(matrix)
    if header_index < 0 or header_index >= len(matrix):
        raise ValueError("表头行超出文件范围")
    headers = normalize_headers(matrix[header_index])
    if not headers:
        raise ValueError("未识别到有效表头")
    rows: list[dict[str, str]] = []
    for values in matrix[header_index + 1 :]:
        row = {
            header: normalize_text(values[index] if index < len(values) else "")
            for index, header in enumerate(headers)
        }
        if any(row.values()):
            rows.append(row)
    return WorksheetData(
        sheet_name=selected_sheet,
        header_row=header_index + 1,
        headers=headers,
        rows=rows,
    )


def read_csv_matrix(file_path: Path) -> list[list[Any]]:
    raw = file_path.read_bytes().decode("utf-8-sig", errors="replace")
    return [list(row) for row in csv.reader(raw.splitlines())]


def read_excel_matrix(file_path: Path, sheet_name: str | None = None) -> tuple[list[list[Any]], str]:
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.worksheets[0]
    return [list(row) for row in sheet.iter_rows(values_only=True)], str(sheet.title)


def detect_header_row(matrix: list[list[Any]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(matrix[:8]):
        cells = [normalize_text(cell) for cell in row]
        score = sum(1 for cell in cells if cell) + sum(3 for cell in cells if match_field(cell))
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def normalize_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}
    for index, value in enumerate(values):
        header = normalize_text(value) or f"未命名列{index + 1}"
        count = used.get(header, 0)
        used[header] = count + 1
        if count:
            header = f"{header}_{count + 1}"
        headers.append(header)
    while headers and headers[-1].startswith("未命名列"):
        headers.pop()
    return headers


def match_field(header: str) -> str | None:
    normalized = normalize_key(header)
    for key, aliases in FIELD_ALIASES.items():
        if normalized in {normalize_key(alias) for alias in aliases}:
            return key
    for key, aliases in FIELD_ALIASES.items():
        if any(normalize_key(alias) in normalized for alias in aliases):
            return key
    return None


def suggest_mapping(headers: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    used: set[str] = set()
    for header in headers:
        key = match_field(header)
        if key and key not in mapping and header not in used:
            mapping[key] = header
            used.add(header)
    return mapping


def validate_mapping(field_mapping: dict[str, str | None], headers: list[str]) -> dict[str, str]:
    valid_keys = {str(field["key"]) for field in STANDARD_FIELDS}
    header_set = set(headers)
    resolved: dict[str, str] = {}
    seen_columns: dict[str, str] = {}
    for key, value in field_mapping.items():
        if key not in valid_keys:
            continue
        column = normalize_text(value)
        if not column:
            continue
        if column not in header_set:
            raise ValueError(f"字段 {field_label(key)} 映射的表格列不存在：{column}")
        if column in seen_columns:
            raise ValueError(
                f"表格列“{column}”被重复映射到{field_label(seen_columns[column])}和{field_label(key)}"
            )
        seen_columns[column] = key
        resolved[key] = column
    missing = [key for key in REQUIRED_FIELDS if not resolved.get(key)]
    if missing:
        raise ValueError(f"必须映射字段：{'、'.join(field_label(key) for key in missing)}")
    return resolved


def field_label(key: str) -> str:
    for field in STANDARD_FIELDS:
        if field.get("key") == key:
            return str(field.get("label") or key)
    return key


def apply_fill_down(values: dict[str, str], carry: dict[str, str]) -> None:
    previous_chapter = carry.get("chapter", "")
    if values.get("chapter"):
        carry["chapter"] = values["chapter"]
        if values["chapter"] != previous_chapter and not values.get("group"):
            carry.pop("group", None)
    elif carry.get("chapter"):
        values["chapter"] = carry["chapter"]

    if values.get("group"):
        carry["group"] = values["group"]
    elif carry.get("group"):
        values["group"] = carry["group"]


def import_knowledge_file(
    path: Path,
    *,
    subject: str = "",
    stage: str = "",
    grade: str = "",
    textbook: str = "",
    mapping: dict[str, str] | None = None,
    field_mapping: dict[str, str | None] | None = None,
    defaults: dict[str, str] | None = None,
    sheet_name: str | None = None,
    header_row: int | None = None,
) -> list[KnowledgePoint]:
    worksheet = read_worksheet(path, sheet_name=sheet_name, header_row=header_row)
    resolved_mapping = validate_mapping(field_mapping or mapping or suggest_mapping(worksheet.headers), worksheet.headers)
    default_values = {
        "subject": subject,
        "stage": stage,
        "grade": grade,
        "textbook": textbook,
        **(defaults or {}),
    }
    default_values = {key: normalize_text(value) for key, value in default_values.items()}
    if not default_values.get("subject"):
        raise ValueError("请先填写基础配置：学科")

    points: list[KnowledgePoint] = []
    carry: dict[str, str] = {}
    for row in worksheet.rows:
        values = {
            "subject": default_values.get("subject", ""),
            "stage": default_values.get("stage", ""),
            "grade": default_values.get("grade", ""),
            "textbook": default_values.get("textbook", ""),
            "chapter": clean_heading(row.get(resolved_mapping["chapter"], "")),
            "group": clean_heading(row.get(resolved_mapping["group"], "")),
            "name": clean_heading(row.get(resolved_mapping["name"], "")),
        }
        if not any(values[key] for key in ("chapter", "group", "name")):
            continue
        apply_fill_down(values, carry)
        if not values["name"]:
            continue
        points.append(make_knowledge_point(values))
    if not points:
        raise ValueError("没有可导入的知识点")
    return points
